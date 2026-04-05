"""
Generate a fully-populated test workbook with random scores.

  • All group-stage goal cells → =RANDBETWEEN(0,4)
  • All bracket goal cells    → =RANDBETWEEN(0,4)
  • All bracket penalty cells → =RANDBETWEEN(3,7)  (slightly wider range to
    keep ties rare but still exercise the penalty path)
  • Best-3rd-place team slots → auto-resolved via the hidden Terceros sheet

Run:
    poetry run python generate_randbetween.py

The output file is written next to this script as:
    quiniela_mundial_2026_randbetween.xlsx
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from src.sheets.bracket import build_bracket
from src.sheets.clasificados import build_clasificados
from src.sheets.group_stage import build_group_stage
from src.sheets.references import build_references
from src.sheets.third_place import build_third_place

OUTPUT_FILE = Path("quiniela_mundial_2026_randbetween.xlsx")


def _is_input_yellow(cell) -> bool:
    """True for cells styled with INPUT_YELLOW (FFD966) fill."""
    try:
        return (
            cell.fill.fill_type == "solid"
            and "FFD966" in cell.fill.fgColor.rgb
        )
    except Exception:
        return False


def _fill_group_scores(ws) -> None:
    """Replace every blank INPUT_YELLOW cell with RANDBETWEEN(0,4)."""
    for row in ws.iter_rows():
        for cell in row:
            if _is_input_yellow(cell) and cell.value is None:
                cell.value = "=RANDBETWEEN(0,4)"


def _fill_bracket_scores(ws, bracket_info: dict) -> None:
    """Fill bracket goal cells and penalty cells with random formulas."""
    for row in ws.iter_rows():
        for cell in row:
            if _is_input_yellow(cell):
                cell.value = "=RANDBETWEEN(0,4)"

    for coord in bracket_info["pen_cells"]:
        ws[coord].value = "=RANDBETWEEN(3,7)"


def main() -> None:
    wb = Workbook()
    del wb["Sheet"]

    print("Building group stage tab…")
    standings_start = build_group_stage(wb)

    print("Building Clasificados tab…")
    qualified_refs = build_clasificados(wb, standings_start)

    print("Building Terceros (best third-place) helper sheet…")
    third_place_refs = build_third_place(wb, standings_start)

    print("Building bracket tab…")
    bracket_info = build_bracket(wb, qualified_refs, third_place_refs)

    print("Building references tab…")
    build_references(wb)

    print("Filling random scores (group stage)…")
    _fill_group_scores(wb["Fase de Grupos"])

    print("Filling random scores (bracket goal cells)…")
    _fill_bracket_scores(wb["Bracket"], bracket_info)

    wb.active = wb["Fase de Grupos"]
    wb.save(OUTPUT_FILE)
    print(f"✅  Saved → {OUTPUT_FILE}")
    print(f"   Penalty cell count: {len(bracket_info['pen_cells'])}")


if __name__ == "__main__":
    main()
