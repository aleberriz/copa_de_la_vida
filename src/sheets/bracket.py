"""
Builds the 'Bracket' worksheet.

Team names in all rounds are driven by Excel formulas:
  • R32 group slots  → reference the Clasificados sheet
  • R32 3rd-place    → auto-resolved via the hidden Terceros sheet
  • R16 onwards      → winner_formula() referencing the previous round's cells
  • Final winner     → winner_formula() on the two SF cells
  • 3rd-place match  → loser_formula() on the two SF cells
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Border, Font
from openpyxl.utils import get_column_letter

from src.data import (
    FINAL, QF_LEFT, QF_RIGHT, R16_LEFT, R16_RIGHT,
    R32_LEFT, R32_RIGHT, SF_LEFT, SF_RIGHT, THIRD,
)
from src.styles import (
    BRACKET_BG, DARK_GREY, FINAL_TEAM_BG, GOLD, MATCH_BG,
    NAVY, PEN_BG, SILVER, WHITE,
    INPUT_BORDER, PEN_BORDER,
    _side, abs_ref, center, fill, input_cell, left_align,
    loser_formula, pen_cell, right_align, winner_formula,
)

# ---------------------------------------------------------------------------
# Column layout — 9 rounds × 4 cols each, no overlaps
# R32-L(1-4) R16-L(5-8) QF-L(9-12) SF-L(13-16)
# FINAL(17-20)
# SF-R(21-24) QF-R(25-28) R16-R(29-32) R32-R(33-36)
# ---------------------------------------------------------------------------
TOTAL_COLS = 36

R32L = (1,  2,  3,  4)
R16L = (5,  6,  7,  8)
QFL  = (9,  10, 11, 12)
SFL  = (13, 14, 15, 16)
FIN  = (17, 18, 19, 20)
SFR  = (21, 22, 23, 24)
QFR  = (25, 26, 27, 28)
R16R = (29, 30, 31, 32)
R32R = (33, 34, 35, 36)

# Row spacing: 4 rows per match slot (label row, match row, pen row, gap)
SLOT = 4
BASE = 8   # first R32 match slot start row


MatchInfo = dict  # keys: tl, tr, gl, gr, pl, pr  (absolute cell refs)


def build_bracket(wb: Workbook,
                  qualified_refs: dict[str, dict[str, str]],
                  third_place_refs: dict[str, str]) -> dict:
    """
    Build the Bracket sheet.

    Parameters
    ----------
    qualified_refs
        Output of build_clasificados — maps group → {'1st': formula, '2nd': formula}.
    third_place_refs
        Output of build_third_place — maps slot codes like '3rd:1E' → formula
        strings referencing the hidden Terceros sheet.

    Returns
    -------
    dict with key:
        'pen_cells' : list[str]  — absolute coords of all penalty score cells.
    """
    ws = wb.create_sheet("Bracket")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 70

    _set_column_widths(ws)

    # Compute row positions
    r32_rows = [BASE + i * SLOT for i in range(8)]
    r16_rows = [(r32_rows[i*2] + r32_rows[i*2+1]) // 2 + 1 for i in range(4)]
    qf_rows  = [(r16_rows[i*2] + r16_rows[i*2+1]) // 2 for i in range(2)]
    sf_row   = (qf_rows[0] + qf_rows[1]) // 2
    final_row = sf_row
    third_row = r32_rows[-1] + SLOT + 3
    max_row   = third_row + SLOT + 4

    # Paint background
    bg = fill(BRACKET_BG)
    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = 18
        for c in range(1, TOTAL_COLS + 1):
            ws.cell(row=r, column=c).fill = bg

    _write_title(ws)
    _write_legend_row(ws)
    _write_round_headers(ws, sf_row, final_row)

    pen_cells: list[str] = []

    # ── Resolve R32 team formulas from qualified_refs / third_place_refs
    def _team_ref(code: str) -> str | None:
        """
        Convert e.g. 'A1' → formula string, '3rd:1E' → formula from Terceros.
        """
        if code.startswith("3rd:"):
            return third_place_refs[code]
        grp, rank = code[0], code[1]
        key = "1st" if rank == "1" else "2nd"
        return qualified_refs[grp][key]

    # ── Draw R32 — left (8 matches) ──────────────────────────────────────
    r32l_info: list[MatchInfo] = []
    for i, (tl_code, tr_code, date, venue) in enumerate(R32_LEFT):
        info = _draw_match(
            ws, r32_rows[i], R32L,
            team_l=_team_ref(tl_code),
            team_r=_team_ref(tr_code),
            date=date, venue=venue,
            _pen_cells_out=pen_cells,
        )
        r32l_info.append(info)

    # ── Draw R32 — right (8 matches) ────────────────────────────────────
    r32r_info: list[MatchInfo] = []
    for i, (tl_code, tr_code, date, venue) in enumerate(R32_RIGHT):
        info = _draw_match(
            ws, r32_rows[i], R32R,
            team_l=_team_ref(tl_code),
            team_r=_team_ref(tr_code),
            date=date, venue=venue,
            _pen_cells_out=pen_cells,
        )
        r32r_info.append(info)

    # ── Draw R16 — left (4 matches, winners of R32-L pairs) ──────────────
    r16l_info: list[MatchInfo] = []
    for i, (a, b, date, venue) in enumerate(R16_LEFT):
        info = _draw_match(
            ws, r16_rows[i], R16L,
            team_l=winner_formula(**r32l_info[a]),
            team_r=winner_formula(**r32l_info[b]),
            date=date, venue=venue,
            _pen_cells_out=pen_cells,
        )
        r16l_info.append(info)

    # ── Draw R16 — right (4 matches, winners of R32-R pairs) ─────────────
    r16r_info: list[MatchInfo] = []
    for i, (a, b, date, venue) in enumerate(R16_RIGHT):
        info = _draw_match(
            ws, r16_rows[i], R16R,
            team_l=winner_formula(**r32r_info[a]),
            team_r=winner_formula(**r32r_info[b]),
            date=date, venue=venue,
            _pen_cells_out=pen_cells,
        )
        r16r_info.append(info)

    # ── Draw QF — left ───────────────────────────────────────────────────
    qfl_info: list[MatchInfo] = []
    for i, (a, b, date, venue) in enumerate(QF_LEFT):
        info = _draw_match(
            ws, qf_rows[i], QFL,
            team_l=winner_formula(**r16l_info[a]),
            team_r=winner_formula(**r16l_info[b]),
            date=date, venue=venue,
            _pen_cells_out=pen_cells,
        )
        qfl_info.append(info)

    # ── Draw QF — right ──────────────────────────────────────────────────
    qfr_info: list[MatchInfo] = []
    for i, (a, b, date, venue) in enumerate(QF_RIGHT):
        info = _draw_match(
            ws, qf_rows[i], QFR,
            team_l=winner_formula(**r16r_info[a]),
            team_r=winner_formula(**r16r_info[b]),
            date=date, venue=venue,
            _pen_cells_out=pen_cells,
        )
        qfr_info.append(info)

    # ── Draw SF — left ───────────────────────────────────────────────────
    a, b, sf_date_l, sf_venue_l = SF_LEFT
    sfl_info = _draw_match(
        ws, sf_row, SFL,
        team_l=winner_formula(**qfl_info[a]),
        team_r=winner_formula(**qfl_info[b]),
        date=sf_date_l, venue=sf_venue_l,
        _pen_cells_out=pen_cells,
    )

    # ── Draw SF — right ──────────────────────────────────────────────────
    a, b, sf_date_r, sf_venue_r = SF_RIGHT
    sfr_info = _draw_match(
        ws, sf_row, SFR,
        team_l=winner_formula(**qfr_info[a]),
        team_r=winner_formula(**qfr_info[b]),
        date=sf_date_r, venue=sf_venue_r,
        _pen_cells_out=pen_cells,
    )

    # ── Draw Final ───────────────────────────────────────────────────────
    fin_date, fin_venue = FINAL
    fin_info = _draw_match(
        ws, final_row, FIN,
        team_l=winner_formula(**sfl_info),
        team_r=winner_formula(**sfr_info),
        date=fin_date, venue=fin_venue,
        is_final=True,
        _pen_cells_out=pen_cells,
    )

    # Champion banner (formula-driven team name)
    champ_row = final_row + 3
    champ_formula = winner_formula(**fin_info)
    ws.merge_cells(start_row=champ_row, start_column=FIN[0],
                   end_row=champ_row, end_column=FIN[3])
    cc = ws.cell(row=champ_row, column=FIN[0],
                 value=f'=IF({fin_info["gl"]}<>"","🏆 "&{champ_formula[1:]}&" 🏆","🏆 WORLD CHAMPION 🏆")')
    cc.font = Font(name="Calibri", size=12, bold=True, color=GOLD)
    cc.fill = fill(NAVY)
    cc.alignment = center()
    cc.border = Border(bottom=_side("medium", GOLD))
    ws.row_dimensions[champ_row].height = 22

    # ── Draw Third-place match ────────────────────────────────────────────
    tp_date, tp_venue = THIRD
    ws.merge_cells(start_row=third_row - 1, start_column=FIN[0],
                   end_row=third_row - 1, end_column=FIN[3])
    t3h = ws.cell(row=third_row - 1, column=FIN[0],
                  value="THIRD PLACE MATCH")
    t3h.font = Font(name="Calibri", size=9, bold=True, color=SILVER)
    t3h.fill = fill(BRACKET_BG)
    t3h.alignment = center()
    t3h.border = Border(bottom=_side("thin", SILVER))
    ws.row_dimensions[third_row - 1].height = 14

    _draw_match(
        ws, third_row, FIN,
        team_l=loser_formula(**sfl_info),
        team_r=loser_formula(**sfr_info),
        date=tp_date, venue=tp_venue,
        _pen_cells_out=pen_cells,
    )

    return {"pen_cells": pen_cells}


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------
def _draw_match(
    ws,
    row: int,
    cols: tuple[int, int, int, int],
    team_l,         # str formula / literal
    team_r,         # str formula / literal
    date: str,
    venue: str,
    is_final: bool = False,
    _pen_cells_out: list | None = None,
) -> MatchInfo:
    """
    Draw a single knockout match (label row + match row + penalty row).

    Returns a MatchInfo dict with absolute cell references for all
    six cells: tl, tr, gl, gr, pl, pr.
    """
    tl_col, gl_col, gr_col, tr_col = cols
    gold_c = GOLD if is_final else "3D5A80"
    tbg    = fill(FINAL_TEAM_BG if is_final else MATCH_BG)

    # Label row (date + venue)
    lbl_row = row
    ws.merge_cells(start_row=lbl_row, start_column=tl_col,
                   end_row=lbl_row, end_column=tr_col)
    dl = ws.cell(row=lbl_row, column=tl_col,
                 value=f"{date}  ·  {venue}")
    dl.font = Font(name="Calibri", size=7, italic=True,
                   color=GOLD if is_final else "607D8B")
    dl.fill = fill(BRACKET_BG)
    dl.alignment = center()
    ws.row_dimensions[lbl_row].height = 13

    # Match row
    mrow = row + 1
    ws.row_dimensions[mrow].height = 24 if is_final else 20

    # Team left
    tl_cell = ws.cell(row=mrow, column=tl_col)
    tl_cell.value = team_l
    tl_cell.fill = tbg
    tl_cell.font = Font(name="Calibri", size=12 if is_final else 9,
                        bold=True, color=GOLD if is_final else WHITE)
    tl_cell.alignment = right_align()
    tl_cell.border = Border(
        top=_side("medium" if is_final else "thin", gold_c),
        bottom=_side("medium" if is_final else "thin", gold_c),
        left=_side("medium", gold_c),
    )

    # Goals
    gl_cell = input_cell(ws, mrow, gl_col, size=14 if is_final else 12)
    gr_cell = input_cell(ws, mrow, gr_col, size=14 if is_final else 12)

    # Team right
    tr_cell = ws.cell(row=mrow, column=tr_col)
    tr_cell.value = team_r
    tr_cell.fill = tbg
    tr_cell.font = Font(name="Calibri", size=12 if is_final else 9,
                        bold=True, color=GOLD if is_final else WHITE)
    tr_cell.alignment = left_align()
    tr_cell.border = Border(
        top=_side("medium" if is_final else "thin", gold_c),
        bottom=_side("medium" if is_final else "thin", gold_c),
        right=_side("medium", gold_c),
    )

    # Penalty row
    prow = row + 2
    ws.row_dimensions[prow].height = 14

    pl_lbl = ws.cell(row=prow, column=tl_col, value="Pen →")
    pl_lbl.font = Font(name="Calibri", size=7, italic=True, color="607D8B")
    pl_lbl.fill = fill(PEN_BG)
    pl_lbl.alignment = right_align()
    pl_lbl.border = PEN_BORDER

    pl_cell = pen_cell(ws, prow, gl_col)
    pr_cell = pen_cell(ws, prow, gr_col)
    if _pen_cells_out is not None:
        _pen_cells_out.extend([pl_cell.coordinate, pr_cell.coordinate])

    pr_lbl = ws.cell(row=prow, column=tr_col, value="← Pen")
    pr_lbl.font = Font(name="Calibri", size=7, italic=True, color="607D8B")
    pr_lbl.fill = fill(PEN_BG)
    pr_lbl.alignment = left_align()
    pr_lbl.border = PEN_BORDER

    return {
        "tl": abs_ref(tl_col, mrow),
        "tr": abs_ref(tr_col, mrow),
        "gl": abs_ref(gl_col, mrow),
        "gr": abs_ref(gr_col, mrow),
        "pl": abs_ref(gl_col, prow),
        "pr": abs_ref(gr_col, prow),
    }


def _write_title(ws) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=TOTAL_COLS)
    t = ws.cell(row=1, column=1,
                value="🏆  FIFA WORLD CUP 2026  ·  KNOCKOUT BRACKET  🏆")
    t.font = Font(name="Calibri", size=20, bold=True, color=GOLD)
    t.fill = fill(BRACKET_BG)
    t.alignment = center()
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 8


def _write_legend_row(ws) -> None:
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=TOTAL_COLS)
    leg = ws.cell(
        row=3, column=1,
        value=(
            "  ✏️ Yellow cells = goals (90 min + ET combined)  ·  "
            "Blue-grey rows = Penalty shootout score (only if tied after ET)  ·  "
            "Best 3rd-place qualifiers are resolved automatically from group-stage results"
        ),
    )
    leg.font = Font(name="Calibri", size=8, italic=True, color="8EAADB")
    leg.fill = fill(BRACKET_BG)
    leg.alignment = left_align()
    ws.row_dimensions[3].height = 14


def _write_round_headers(ws, sf_row: int, final_row: int) -> None:
    hdr_row = 6
    ws.row_dimensions[hdr_row].height = 26

    round_specs = [
        (R32L, "ROUND OF 32",    "Jun 28 – Jul 3"),
        (R16L, "ROUND OF 16",    "Jul 4 – 7"),
        (QFL,  "QUARTER-FINALS", "Jul 9 – 11"),
        (SFL,  "SEMI-FINALS",    "Jul 14 – 15"),
        (FIN,  "🏆  FINAL  🏆", "Jul 19 · New York / NJ"),
        (SFR,  "SEMI-FINALS",    "Jul 14 – 15"),
        (QFR,  "QUARTER-FINALS", "Jul 9 – 11"),
        (R16R, "ROUND OF 16",    "Jul 4 – 7"),
        (R32R, "ROUND OF 32",    "Jun 28 – Jul 3"),
    ]
    for cols, round_name, dates in round_specs:
        sc_col, _, _, ec_col = cols
        is_final = "FINAL" in round_name
        ws.merge_cells(start_row=hdr_row, start_column=sc_col,
                       end_row=hdr_row, end_column=ec_col)
        c = ws.cell(row=hdr_row, column=sc_col,
                    value=f"{round_name}\n{dates}")
        c.font = Font(name="Calibri", size=9, bold=True,
                      color=GOLD if is_final else WHITE)
        c.fill = fill(NAVY if is_final else BRACKET_BG)
        c.alignment = center(wrap=True)
        c.border = Border(
            bottom=_side("medium", GOLD if is_final else "3D5A80")
        )


def _set_column_widths(ws) -> None:
    for c in range(1, TOTAL_COLS + 1):
        ltr = get_column_letter(c)
        # cols 1,4,5,8,9,12,13,16,17,20,21,24,25,28,29,32,33,36 = team name cols
        if (c - 1) % 4 in (0, 3):
            ws.column_dimensions[ltr].width = 14
        else:
            ws.column_dimensions[ltr].width = 5
