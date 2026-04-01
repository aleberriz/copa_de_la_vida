"""Builds the 'Fase de Grupos' worksheet."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Border, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.data import GROUP_MATCHES, GROUPS
from src.styles import (
    AMBER_FONT, AMBER_QUAL, DARK_GREY, GREEN_FONT, GREEN_QUAL,
    GREEN_QUAL_2, INPUT_YELLOW, LIGHT_BLUE, NAVY, RED_ELIM, RED_FONT,
    SILVER, THIN_BORDER, WHITE,
    _side, center, fill, input_cell, left_align, sc,
)

# Column layout (1-indexed)
COL_MD     = 1
COL_DATE   = 2
COL_HOME   = 3
COL_GL     = 4   # home goals — user input
COL_GR     = 5   # away goals — user input
COL_AWAY   = 6
COL_VENUE  = 7
COL_SEP    = 8   # spacer
COL_POS    = 9
COL_TEAM   = 10
COL_GP     = 11
COL_W      = 12
COL_D      = 13
COL_L      = 14
COL_GF     = 15
COL_GA     = 16
COL_GD     = 17
COL_PTS    = 18
COL_SORTKEY = 19  # hidden tiebreaker column (Pts→GD→GF→position)

EASTER_EGGS = {
    "A": "  ·  ¡Salinator por Chequia! 🇨🇿",
    "I": "  ·  Allez les bleus! 🇫🇷",
}


def build_group_stage(wb: Workbook) -> dict[str, int]:
    """
    Build the group-stage sheet.

    Returns
    -------
    standings_start : dict[group_letter, first_standings_row]
        Row number of the first team in each group's standings table,
        needed by the Clasificados tab to build ranking formulas.
    """
    ws = wb.create_sheet("Fase de Grupos")
    ws.sheet_view.showGridLines = False

    _set_column_widths(ws)

    row = 1
    standings_start: dict[str, int] = {}

    for grp in GROUPS:
        teams   = GROUPS[grp]
        matches = GROUP_MATCHES[grp]
        easter  = EASTER_EGGS.get(grp, "")

        # ── Group header ────────────────────────────────────────────────
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=COL_SORTKEY)
        hdr = ws.cell(row=row, column=1,
                      value=f"  GROUP {grp}{easter}")
        hdr.font = Font(name="Calibri", size=14, bold=True,
                        color=WHITE, italic=bool(easter))
        hdr.fill = fill(NAVY)
        hdr.alignment = left_align()
        hdr.border = Border(
            left=_side("medium", "F4C842"), right=_side("medium", "F4C842"),
            top=_side("medium", "F4C842"), bottom=_side("medium", "F4C842"),
        )
        ws.row_dimensions[row].height = 24
        row += 1

        # ── Standings table header ───────────────────────────────────────
        hdr_row = row
        for col_off, label in enumerate(
            ["#", "Team", "GP", "W", "D", "L", "GF", "GA", "GD", "Pts"],
            start=0,
        ):
            c = ws.cell(row=row, column=COL_POS + col_off, value=label)
            c.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
            c.fill = fill(DARK_GREY)
            c.alignment = center()
            c.border = THIN_BORDER
        ws.row_dimensions[row].height = 16
        row += 1

        # ── Standings rows ───────────────────────────────────────────────
        s_start = row
        standings_start[grp] = s_start
        s_rows: list[int] = []

        for team in teams:
            sc(ws, row, COL_POS, "",
               font=Font(name="Calibri", size=9, bold=True),
               fill_=fill(SILVER), border=THIN_BORDER, align=center())
            sc(ws, row, COL_TEAM, team,
               font=Font(name="Calibri", size=10, bold=True),
               fill_=fill(WHITE), border=THIN_BORDER, align=left_align())
            for col in range(COL_GP, COL_SORTKEY):
                sc(ws, row, col, 0,
                   font=Font(name="Calibri", size=10),
                   fill_=fill(WHITE), border=THIN_BORDER, align=center())
            # Sort-key column: hidden
            ws.cell(row=row, column=COL_SORTKEY)   # will be written below
            ws.row_dimensions[row].height = 17
            s_rows.append(row)
            row += 1

        row += 1  # spacer

        # ── Match schedule header ────────────────────────────────────────
        for col_off, label in enumerate(
            ["MD", "Date / Time (ET)", "Home Team", "G", "G", "Away Team", "Venue"],
            start=0,
        ):
            c = ws.cell(row=row, column=1 + col_off, value=label)
            c.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
            c.fill = fill("3D5A80")
            c.alignment = center()
            c.border = THIN_BORDER
        ws.row_dimensions[row].height = 16
        row += 1

        # ── Match rows (user input cells) ────────────────────────────────
        match_refs: list[dict] = []
        for md, date, time_et, home, away, venue in matches:
            sc(ws, row, COL_MD, md,
               font=Font(name="Calibri", size=9),
               fill_=fill(LIGHT_BLUE), border=THIN_BORDER, align=center())
            sc(ws, row, COL_DATE, f"{date} · {time_et}",
               font=Font(name="Calibri", size=8),
               fill_=fill(LIGHT_BLUE), border=THIN_BORDER, align=center())
            sc(ws, row, COL_HOME, home,
               font=Font(name="Calibri", size=10, bold=True),
               fill_=fill(LIGHT_BLUE), border=THIN_BORDER, align=left_align())
            hg = input_cell(ws, row, COL_GL)
            ag = input_cell(ws, row, COL_GR)
            sc(ws, row, COL_AWAY, away,
               font=Font(name="Calibri", size=10, bold=True),
               fill_=fill(LIGHT_BLUE), border=THIN_BORDER, align=left_align())
            sc(ws, row, COL_VENUE, venue,
               font=Font(name="Calibri", size=8),
               fill_=fill(LIGHT_BLUE), border=THIN_BORDER, align=left_align())
            ws.row_dimensions[row].height = 20
            match_refs.append({
                "home": home, "away": away,
                "hg": hg.coordinate, "ag": ag.coordinate,
            })
            row += 1

        # ── Write standings formulas ─────────────────────────────────────
        for i, srow in enumerate(s_rows):
            team  = teams[i]
            hm    = [(m["hg"], m["ag"]) for m in match_refs if m["home"] == team]
            aw    = [(m["hg"], m["ag"]) for m in match_refs if m["away"] == team]
            all_m = hm + aw

            def _played(pairs):
                return "+".join(
                    f"IF(AND({h}<>\"\",{a}<>\"\"),1,0)" for h, a in pairs
                ) or "0"

            def _sum_cond(pairs, pick_h, pick_a):
                parts = [
                    f"IF(AND({h}<>\"\",{a}<>\"\"),{pick_h(h,a)},0)"
                    for h, a in pairs
                ]
                return "+".join(parts) or "0"

            gp  = f"={_played(all_m)}"
            gf  = "=" + "+".join(filter(None, [
                _sum_cond(hm,  lambda h,a: h, lambda h,a: a),
                _sum_cond(aw,  lambda h,a: a, lambda h,a: h),
            ])) if (hm or aw) else "=0"
            ga  = "=" + "+".join(filter(None, [
                _sum_cond(hm,  lambda h,a: a, lambda h,a: h),
                _sum_cond(aw,  lambda h,a: h, lambda h,a: a),
            ])) if (hm or aw) else "=0"
            wins = "=" + "+".join(filter(None, [
                "+".join(f"IF(AND({h}<>\"\",{a}<>\"\",{h}>{a}),1,0)" for h,a in hm),
                "+".join(f"IF(AND({h}<>\"\",{a}<>\"\",{a}>{h}),1,0)" for h,a in aw),
            ])) if (hm or aw) else "=0"
            draws = "=" + "+".join(
                f"IF(AND({h}<>\"\",{a}<>\"\",{h}={a}),1,0)"
                for h, a in all_m
            ) if all_m else "=0"
            losses = "=" + "+".join(filter(None, [
                "+".join(f"IF(AND({h}<>\"\",{a}<>\"\",{h}<{a}),1,0)" for h,a in hm),
                "+".join(f"IF(AND({h}<>\"\",{a}<>\"\",{a}<{h}),1,0)" for h,a in aw),
            ])) if (hm or aw) else "=0"

            ws.cell(row=srow, column=COL_GP).value  = gp
            ws.cell(row=srow, column=COL_W).value   = wins
            ws.cell(row=srow, column=COL_D).value   = draws
            ws.cell(row=srow, column=COL_L).value   = losses
            ws.cell(row=srow, column=COL_GF).value  = gf
            ws.cell(row=srow, column=COL_GA).value  = ga

            gf_r = ws.cell(row=srow, column=COL_GF).coordinate
            ga_r = ws.cell(row=srow, column=COL_GA).coordinate
            w_r  = ws.cell(row=srow, column=COL_W).coordinate
            d_r  = ws.cell(row=srow, column=COL_D).coordinate
            ws.cell(row=srow, column=COL_GD).value  = f"={gf_r}-{ga_r}"
            ws.cell(row=srow, column=COL_PTS).value = f"=3*{w_r}+{d_r}"

            # Sort key: Pts * 10M + (GD+99) * 100K + GF * 100 + position_tiebreak
            pts_r = ws.cell(row=srow, column=COL_PTS).coordinate
            gd_r  = ws.cell(row=srow, column=COL_GD).coordinate
            gff_r = ws.cell(row=srow, column=COL_GF).coordinate
            tiebreak = 4 - i  # 4,3,2,1 → ensures unique keys within group
            ws.cell(row=srow, column=COL_SORTKEY).value = (
                f"={pts_r}*10000000+({gd_r}+99)*100000+{gff_r}*100+{tiebreak}"
            )

        # ── Conditional formatting on standings ──────────────────────────
        q_range = (
            f"{get_column_letter(COL_POS)}{s_start}:"
            f"{get_column_letter(COL_PTS)}{s_start + 3}"
        )
        pts_letter = get_column_letter(COL_PTS)

        def _cf(rank: int):
            return [
                f"RANK(${pts_letter}{s_start},"
                f"${pts_letter}${s_start}:${pts_letter}${s_start + 3},0)={rank}"
            ]

        ws.conditional_formatting.add(q_range, FormulaRule(
            formula=_cf(1), fill=PatternFill("solid", fgColor=GREEN_QUAL),
            font=Font(color=GREEN_FONT, bold=True),
        ))
        ws.conditional_formatting.add(q_range, FormulaRule(
            formula=_cf(2), fill=PatternFill("solid", fgColor=GREEN_QUAL_2),
            font=Font(color=GREEN_FONT),
        ))
        ws.conditional_formatting.add(q_range, FormulaRule(
            formula=_cf(3), fill=PatternFill("solid", fgColor=AMBER_QUAL),
            font=Font(color=AMBER_FONT),
        ))
        ws.conditional_formatting.add(q_range, FormulaRule(
            formula=_cf(4), fill=PatternFill("solid", fgColor=RED_ELIM),
            font=Font(color=RED_FONT),
        ))

        row += 2  # gap between groups

    # ── Legend ───────────────────────────────────────────────────────────
    _write_legend(ws, row)

    # ── Hide sort-key column ─────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(COL_SORTKEY)].hidden = True

    return standings_start


def _set_column_widths(ws) -> None:
    widths = {
        COL_MD:      4,
        COL_DATE:   16,
        COL_HOME:   22,
        COL_GL:      5,
        COL_GR:      5,
        COL_AWAY:   22,
        COL_VENUE:  20,
        COL_SEP:     2,
        COL_POS:     3,
        COL_TEAM:   22,
        COL_GP:      4,
        COL_W:       4,
        COL_D:       4,
        COL_L:       4,
        COL_GF:      4,
        COL_GA:      4,
        COL_GD:      5,
        COL_PTS:     5,
        COL_SORTKEY: 0,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_legend(ws, row: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    sc(ws, row, 1, "  LEGEND",
       font=Font(name="Calibri", size=10, bold=True, color=WHITE),
       fill_=fill(DARK_GREY), border=THIN_BORDER, align=left_align())
    ws.row_dimensions[row].height = 18
    row += 1
    for bg, fc, label in [
        (GREEN_QUAL,   GREEN_FONT, "✅  1st & 2nd place — qualify automatically to Round of 32"),
        (AMBER_QUAL,   AMBER_FONT, "⚠️   3rd place — may qualify as one of the 8 best third-place finishers"),
        (RED_ELIM,     RED_FONT,   "❌  4th place — eliminated"),
        (INPUT_YELLOW, DARK_GREY,  "✏️   Yellow cell — enter number of goals scored (integer ≥ 0)"),
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value=f"  {label}")
        c.font = Font(name="Calibri", size=9, bold=True, color=fc)
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = THIN_BORDER
        c.alignment = left_align()
        ws.row_dimensions[row].height = 16
        row += 1
