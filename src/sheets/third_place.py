"""
Builds the hidden 'Terceros' worksheet — ranks all 12 third-place
finishers, determines which 8 advance, and resolves the FIFA Annex C
bracket assignments via a 495-row lookup table.

The bracket tab references the 8 result cells produced here instead of
requiring manual entry of best-third-place qualifiers.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.data import (
    GROUP_POWER, GROUPS, THIRD_PLACE_COMBINATIONS, THIRD_PLACE_SLOT_ORDER,
)
from src.sheets.group_stage import COL_GD, COL_GF, COL_PTS, COL_SORTKEY, COL_TEAM
from src.styles import abs_ref

SHEET_NAME = "Terceros"
GS_SHEET = "Fase de Grupos"

# Layout — stats section
_COL_GRP = 1   # A  group letter
_COL_TEAM = 2  # B  3rd-place team name (formula)
_COL_PTS = 3   # C  points
_COL_GD = 4    # D  goal difference
_COL_GF = 5    # E  goals for
_COL_SK = 6    # F  cross-group sort key
_COL_RANK = 7  # G  rank (1-12)
_COL_QUAL = 8  # H  qualifies? (1/0)
_COL_POW = 9   # I  power-of-2 value
_COL_CONTRIB = 10  # J  contribution to combo key (QUAL * POW)

STATS_HEADER_ROW = 1
STATS_FIRST_ROW = 2  # group A
STATS_LAST_ROW = 13   # group L

COMBO_KEY_ROW = 15

# Lookup table section
LUT_COL_KEY = 12    # L  combination key
LUT_COL_FIRST = 13  # M  vs_1A  (first assignment column)
LUT_FIRST_ROW = 2
LUT_LAST_ROW = LUT_FIRST_ROW + len(THIRD_PLACE_COMBINATIONS) - 1  # 496

# Result section — 8 cells that resolve to team names
RESULT_HEADER_ROW = 17
RESULT_FIRST_ROW = 18


def build_third_place(
    wb: Workbook,
    standings_start: dict[str, int],
) -> dict[str, str]:
    """
    Build the hidden Terceros helper sheet.

    Parameters
    ----------
    standings_start
        Maps group letter → first row of that group's standings table
        in the 'Fase de Grupos' sheet.

    Returns
    -------
    third_place_refs
        Maps slot codes ('3rd:1A', '3rd:1B', …) → Excel formula strings
        ready to set as cell values in the Bracket sheet.
    """
    ws = wb.create_sheet(SHEET_NAME)
    ws.sheet_view.showGridLines = False
    ws.sheet_state = "hidden"

    _write_stats_section(ws, standings_start)
    _write_combo_key(ws)
    _write_lookup_table(ws)
    return _write_results(ws)


def _write_stats_section(ws, standings_start: dict[str, int]) -> None:
    """One row per group with 3rd-place team stats and ranking."""

    group_list = list(GROUPS.keys())
    sk_col_gs = get_column_letter(COL_SORTKEY)
    team_col_gs = get_column_letter(COL_TEAM)
    pts_col_gs = get_column_letter(COL_PTS)
    gd_col_gs = get_column_letter(COL_GD)
    gf_col_gs = get_column_letter(COL_GF)

    sk_col = get_column_letter(_COL_SK)
    rank_col = get_column_letter(_COL_RANK)
    sk_range_abs = (
        f"${sk_col}${STATS_FIRST_ROW}:${sk_col}${STATS_LAST_ROW}"
    )

    for i, grp in enumerate(group_list):
        row = STATS_FIRST_ROW + i
        s = standings_start[grp]

        gs_sk_range = f"'{GS_SHEET}'!${sk_col_gs}${s}:${sk_col_gs}${s + 3}"
        gs_team_range = f"'{GS_SHEET}'!${team_col_gs}${s}:${team_col_gs}${s + 3}"
        gs_pts_range = f"'{GS_SHEET}'!${pts_col_gs}${s}:${pts_col_gs}${s + 3}"
        gs_gd_range = f"'{GS_SHEET}'!${gd_col_gs}${s}:${gd_col_gs}${s + 3}"
        gs_gf_range = f"'{GS_SHEET}'!${gf_col_gs}${s}:${gf_col_gs}${s + 3}"

        third_pos = f"MATCH(LARGE({gs_sk_range},3),{gs_sk_range},0)"

        ws.cell(row=row, column=_COL_GRP, value=grp)
        ws.cell(row=row, column=_COL_TEAM,
                value=f"=IFERROR(INDEX({gs_team_range},{third_pos}),\"\")")
        ws.cell(row=row, column=_COL_PTS,
                value=f"=IFERROR(INDEX({gs_pts_range},{third_pos}),0)")
        ws.cell(row=row, column=_COL_GD,
                value=f"=IFERROR(INDEX({gs_gd_range},{third_pos}),0)")
        ws.cell(row=row, column=_COL_GF,
                value=f"=IFERROR(INDEX({gs_gf_range},{third_pos}),0)")

        pts_ref = abs_ref(_COL_PTS, row)
        gd_ref = abs_ref(_COL_GD, row)
        gf_ref = abs_ref(_COL_GF, row)
        tiebreak = 12 - i
        ws.cell(row=row, column=_COL_SK,
                value=f"={pts_ref}*10000000+({gd_ref}+99)*100000+{gf_ref}*100+{tiebreak}")

        sk_ref = abs_ref(_COL_SK, row)
        ws.cell(row=row, column=_COL_RANK,
                value=f"=RANK({sk_ref},{sk_range_abs},0)")

        rank_ref = abs_ref(_COL_RANK, row)
        ws.cell(row=row, column=_COL_QUAL,
                value=f"=IF({rank_ref}<=8,1,0)")

        ws.cell(row=row, column=_COL_POW, value=GROUP_POWER[grp])

        qual_ref = abs_ref(_COL_QUAL, row)
        pow_ref = abs_ref(_COL_POW, row)
        ws.cell(row=row, column=_COL_CONTRIB,
                value=f"={qual_ref}*{pow_ref}")


def _write_combo_key(ws) -> None:
    """Sum of power-of-2 contributions for the 8 qualifying groups."""
    contrib_col = get_column_letter(_COL_CONTRIB)
    ws.cell(
        row=COMBO_KEY_ROW,
        column=_COL_CONTRIB,
        value=(
            f"=SUM(${contrib_col}${STATS_FIRST_ROW}"
            f":${contrib_col}${STATS_LAST_ROW})"
        ),
    )


def _write_lookup_table(ws) -> None:
    """Write the 495-row FIFA Annex C combination table."""
    for i, combo in enumerate(THIRD_PLACE_COMBINATIONS):
        row = LUT_FIRST_ROW + i
        key = combo[0]
        assignments = combo[1:]  # 8 group letters

        ws.cell(row=row, column=LUT_COL_KEY, value=key)
        for j, grp_letter in enumerate(assignments):
            ws.cell(row=row, column=LUT_COL_FIRST + j, value=grp_letter)


def _write_results(ws) -> dict[str, str]:
    """
    Resolve each of the 8 R32 third-place slots to a team name.

    Uses VLOOKUP on the combo key to find the matching row, then
    INDEX/MATCH from the group letter back to the team name.
    """
    combo_key_ref = abs_ref(_COL_CONTRIB, COMBO_KEY_ROW)
    key_col_letter = get_column_letter(LUT_COL_KEY)
    last_assign_col = LUT_COL_FIRST + 7
    lut_range = (
        f"${key_col_letter}${LUT_FIRST_ROW}"
        f":{get_column_letter(last_assign_col)}${LUT_LAST_ROW}"
    )

    grp_col_letter = get_column_letter(_COL_GRP)
    team_col_letter = get_column_letter(_COL_TEAM)
    grp_range = (
        f"${grp_col_letter}${STATS_FIRST_ROW}"
        f":${grp_col_letter}${STATS_LAST_ROW}"
    )
    team_range = (
        f"${team_col_letter}${STATS_FIRST_ROW}"
        f":${team_col_letter}${STATS_LAST_ROW}"
    )

    third_place_refs: dict[str, str] = {}

    for i, slot_code in enumerate(THIRD_PLACE_SLOT_ORDER):
        row = RESULT_FIRST_ROW + i
        vlookup_col_idx = 2 + i  # column offset within the VLOOKUP range

        ws.cell(row=row, column=_COL_GRP, value=slot_code)

        group_letter_formula = (
            f"VLOOKUP({combo_key_ref},{lut_range},{vlookup_col_idx},0)"
        )
        team_name_formula = (
            f'=IFERROR(INDEX({team_range},'
            f'MATCH({group_letter_formula},{grp_range},0)),"—")'
        )
        ws.cell(row=row, column=_COL_TEAM, value=team_name_formula)

        ref_str = (
            f"='{SHEET_NAME}'!"
            f"${team_col_letter}${row}"
        )
        third_place_refs[f"3rd:{slot_code}"] = ref_str

    return third_place_refs
