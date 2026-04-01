"""
Builds the 'Clasificados' worksheet — a clean summary of which teams
qualified from each group, driven entirely by formulas from the
'Fase de Grupos' sheet.

Also returns a lookup dict used by the bracket tab to reference
first- and second-place teams by formula.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.data import GROUPS
from src.sheets.group_stage import COL_SORTKEY, COL_TEAM
from src.styles import (
    AMBER_FONT, AMBER_QUAL, DARK_GREY, GOLD, GREEN_FONT, GREEN_QUAL,
    GREEN_QUAL_2, NAVY, RED_ELIM, RED_FONT, SILVER, THIN_BORDER, WHITE,
    center, fill, left_align, sc,
)

SHEET_NAME = "Clasificados"
GS_SHEET   = "Fase de Grupos"

# Columns in the Clasificados sheet
COL_GRP   = 1
COL_1ST   = 2
COL_2ND   = 3
COL_3RD   = 4
COL_4TH   = 5

# Data rows: row 4 = Group A, row 5 = Group B, … row 15 = Group L
DATA_ROW_OFFSET = 4   # first data row


def build_clasificados(wb: Workbook,
                       standings_start: dict[str, int]) -> dict[str, dict[str, str]]:
    """
    Build the Clasificados sheet.

    Parameters
    ----------
    standings_start
        Maps group letter → first row of that group's standings table
        in the 'Fase de Grupos' sheet.

    Returns
    -------
    qualified_refs
        Maps group letter → {'1st': formula_str, '2nd': formula_str}
        where formula_str is an Excel formula (e.g. ``='Clasificados'!$B$4``)
        ready to be set as a cell value in another sheet.
    """
    ws = wb.create_sheet(SHEET_NAME)
    ws.sheet_view.showGridLines = False

    for c, w in {COL_GRP: 10, COL_1ST: 22, COL_2ND: 22,
                 COL_3RD: 22, COL_4TH: 22}.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── Title ────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    t = ws.cell(row=1, column=1,
                value="  🏟️  Clasificados · Qualified Teams")
    t.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = left_align()
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    sub = ws.cell(
        row=2, column=1,
        value="  Rankings update automatically as you enter scores in the Fase de Grupos tab.",
    )
    sub.font = Font(name="Calibri", size=9, italic=True, color=DARK_GREY)
    sub.fill = fill(SILVER)
    sub.alignment = left_align()
    ws.row_dimensions[2].height = 16

    # ── Table header ─────────────────────────────────────────────────────
    row = 3
    for col_off, label in enumerate(
        ["Group", "1st Place ✅", "2nd Place ✅", "3rd Place ⚠️", "4th Place ❌"],
        start=0,
    ):
        c = ws.cell(row=row, column=1 + col_off, value=label)
        c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.fill = fill(DARK_GREY)
        c.alignment = center()
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 18

    # ── One row per group ────────────────────────────────────────────────
    qualified_refs: dict[str, dict[str, str]] = {}

    for g_idx, grp in enumerate(GROUPS):
        row = DATA_ROW_OFFSET + g_idx
        s   = standings_start[grp]   # first standings row in group stage sheet

        # Sort-key range in group-stage sheet
        sk_col   = get_column_letter(COL_SORTKEY)
        team_col = get_column_letter(COL_TEAM)
        sk_range = (
            f"'{GS_SHEET}'!${sk_col}${s}:${sk_col}${s + 3}"
        )
        team_range = (
            f"'{GS_SHEET}'!${team_col}${s}:${team_col}${s + 3}"
        )

        # Group label
        sc(ws, row, COL_GRP, f"Group {grp}",
           font=Font(name="Calibri", size=10, bold=True, color=WHITE),
           fill_=fill(NAVY), border=THIN_BORDER, align=center())

        row_bg = fill("F2F2F2") if g_idx % 2 == 0 else fill("FFFFFF")

        for rank, col, bg, fc in [
            (1, COL_1ST, GREEN_QUAL,   GREEN_FONT),
            (2, COL_2ND, GREEN_QUAL_2, GREEN_FONT),
            (3, COL_3RD, AMBER_QUAL,   AMBER_FONT),
            (4, COL_4TH, RED_ELIM,     RED_FONT),
        ]:
            formula = (
                f"=IFERROR(INDEX({team_range},"
                f"MATCH(LARGE({sk_range},{rank}),{sk_range},0)),\"—\")"
            )
            c = ws.cell(row=row, column=col, value=formula)
            c.font = Font(name="Calibri", size=10, bold=True, color=fc)
            c.fill = fill(bg)
            c.border = THIN_BORDER
            c.alignment = center()

        ws.row_dimensions[row].height = 20

        # Build reference strings for use in the Bracket tab
        c1st = get_column_letter(COL_1ST)
        c2nd = get_column_letter(COL_2ND)
        data_row = DATA_ROW_OFFSET + g_idx
        qualified_refs[grp] = {
            "1st": f"='{SHEET_NAME}'!${c1st}${data_row}",
            "2nd": f"='{SHEET_NAME}'!${c2nd}${data_row}",
        }

    # ── Footer note ──────────────────────────────────────────────────────
    note_row = DATA_ROW_OFFSET + len(GROUPS) + 1
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=5)
    n = ws.cell(
        row=note_row, column=1,
        value=(
            "  ⚠️  The 8 best 3rd-place finishers also advance to the Round of 32. "
            "Those slots are determined after all groups complete and must be "
            "entered manually in the Bracket tab."
        ),
    )
    n.font = Font(name="Calibri", size=9, italic=True, color=AMBER_FONT)
    n.fill = fill(AMBER_QUAL)
    n.alignment = left_align(wrap=True)
    n.border = THIN_BORDER
    ws.row_dimensions[note_row].height = 36

    return qualified_refs
