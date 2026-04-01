"""Shared style constants and helper functions for openpyxl."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------
NAVY          = "1F3864"
GOLD          = "F4C842"
WHITE         = "FFFFFF"
DARK_GREY     = "2F2F2F"
SILVER        = "D9D9D9"
INPUT_YELLOW  = "FFD966"
LIGHT_BLUE    = "DEEAF1"

GREEN_QUAL      = "C6EFCE"
GREEN_QUAL_2    = "E2EFDA"
GREEN_FONT      = "276221"
AMBER_QUAL      = "FFEB9C"
AMBER_FONT      = "9C5700"
RED_ELIM        = "FFC7CE"
RED_FONT        = "9C0006"

BRACKET_BG      = "0D1B2A"
MATCH_BG        = "1A2D45"
FINAL_TEAM_BG   = "1A3A0A"
PEN_BG          = "263D5A"
CLASI_BG        = "162032"

# ---------------------------------------------------------------------------
# Borders
# ---------------------------------------------------------------------------
def _side(style: str = "thin", color: str = "000000") -> Side:
    return Side(style=style, color=color)

THIN_BORDER = Border(left=_side(), right=_side(), top=_side(), bottom=_side())

INPUT_BORDER = Border(
    left=_side("medium", GOLD), right=_side("medium", GOLD),
    top=_side("medium", GOLD),  bottom=_side("medium", GOLD),
)
PEN_BORDER = Border(
    left=_side("thin", "607D8B"), right=_side("thin", "607D8B"),
    top=_side("thin", "607D8B"),  bottom=_side("thin", "607D8B"),
)
GOLD_BORDER = Border(
    left=_side("medium", GOLD), right=_side("medium", GOLD),
    top=_side("medium", GOLD),  bottom=_side("medium", GOLD),
)

# ---------------------------------------------------------------------------
# Fills / fonts / alignments
# ---------------------------------------------------------------------------
def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def center(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def right_align() -> Alignment:
    return Alignment(horizontal="right", vertical="center")

# ---------------------------------------------------------------------------
# Cell writer
# ---------------------------------------------------------------------------
def sc(ws, row: int, col: int, value=None, *,
       font=None, fill_=None, border=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font = font
    if fill_:  cell.fill = fill_
    if border: cell.border = border
    if align:  cell.alignment = align
    return cell

# ---------------------------------------------------------------------------
# Data validation (non-negative integer)
# ---------------------------------------------------------------------------
def make_dv(sqref: str) -> DataValidation:
    dv = DataValidation(
        type="whole", operator="greaterThanOrEqual", formula1=0,
        showErrorMessage=True,
        errorTitle="Valor inválido",
        error="Introduce un número entero ≥ 0 (goles no pueden ser negativos).",
        showInputMessage=True,
        promptTitle="Goles",
        prompt="Introduce el número de goles (entero ≥ 0).",
    )
    dv.sqref = sqref
    return dv

def input_cell(ws, row: int, col: int, size: int = 12):
    """Style a cell as a user-input goal cell and attach validation."""
    cell = ws.cell(row=row, column=col)
    cell.fill = fill(INPUT_YELLOW)
    cell.font = Font(name="Calibri", size=size, bold=True, color=DARK_GREY)
    cell.alignment = center()
    cell.border = INPUT_BORDER
    ws.add_data_validation(make_dv(cell.coordinate))
    return cell

def pen_cell(ws, row: int, col: int):
    """Style a cell as an optional penalty-score input."""
    cell = ws.cell(row=row, column=col)
    cell.fill = fill(PEN_BG)
    cell.font = Font(name="Calibri", size=9, italic=True, color="8EAADB")
    cell.alignment = center()
    cell.border = PEN_BORDER
    ws.add_data_validation(make_dv(cell.coordinate))
    return cell

# ---------------------------------------------------------------------------
# Bracket formula helpers
# ---------------------------------------------------------------------------
def abs_ref(col: int, row: int) -> str:
    """Return an absolute Excel cell reference string, e.g. '$B$5'."""
    return f"${get_column_letter(col)}${row}"

def winner_formula(tl: str, gl: str, gr: str, tr: str,
                   pl: str, pr: str) -> str:
    """
    Excel formula that resolves to the winner of a knockout match.
    All args are absolute cell reference strings (e.g. '$A$5').
    Returns '—' when no scores entered yet; 'TBD' if tied with no penalties.
    """
    return (
        f'=IF(AND({gl}<>"",{gr}<>""),'
        f'IF({gl}>{gr},{tl},'
        f'IF({gr}>{gl},{tr},'
        f'IF(AND({pl}<>"",{pr}<>""),'
        f'IF({pl}>{pr},{tl},{tr}),'
        f'"TBD"))),"—")'
    )

def loser_formula(tl: str, gl: str, gr: str, tr: str,
                  pl: str, pr: str) -> str:
    """Excel formula that resolves to the loser of a knockout match."""
    return (
        f'=IF(AND({gl}<>"",{gr}<>""),'
        f'IF({gl}<{gr},{tl},'
        f'IF({gr}<{gl},{tr},'
        f'IF(AND({pl}<>"",{pr}<>""),'
        f'IF({pl}<{pr},{tl},{tr}),'
        f'"TBD"))),"—")'
    )

def column_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width
