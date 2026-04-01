"""
FIFA World Cup 2026 — Quiniela / Football Pool generator.
Run:  poetry run python generate_quiniela.py
Output: quiniela_mundial_2026.xlsx
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE = Path("quiniela_mundial_2026.xlsx")

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
NAVY        = "1F3864"
GOLD        = "F4C842"
WHITE       = "FFFFFF"
DARK_GREY   = "2F2F2F"
INPUT_YELLOW = "FFD966"
GREEN_QUAL  = "C6EFCE"
GREEN_QUAL_FONT = "276221"
AMBER_QUAL  = "FFEB9C"
AMBER_QUAL_FONT = "9C5700"
RED_ELIM    = "FFC7CE"
RED_ELIM_FONT = "9C0006"
SILVER      = "D9D9D9"
BRACKET_BG  = "0D1B2A"
MATCH_BG    = "1A2D45"
PEN_BG      = "263D5A"
LIGHT_BLUE  = "DEEAF1"
FINAL_TEAM_BG = "1A3A0A"

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _side(style="thin", color="000000") -> Side:
    return Side(style=style, color=color)

THIN_BORDER = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
INPUT_BORDER = Border(
    left=_side("medium", GOLD), right=_side("medium", GOLD),
    top=_side("medium", GOLD), bottom=_side("medium", GOLD),
)
PEN_BORDER = Border(
    left=_side("thin", "607D8B"), right=_side("thin", "607D8B"),
    top=_side("thin", "607D8B"), bottom=_side("thin", "607D8B"),
)

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def center(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def right_align() -> Alignment:
    return Alignment(horizontal="right", vertical="center")

def sc(ws, row, col, value=None, *, font=None, fill_=None,
       border=None, align=None) -> object:
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font = font
    if fill_:  cell.fill = fill_
    if border: cell.border = border
    if align:  cell.alignment = align
    return cell

def make_dv(sqref: str) -> DataValidation:
    dv = DataValidation(
        type="whole", operator="greaterThanOrEqual", formula1=0,
        showErrorMessage=True,
        errorTitle="Valor inválido",
        error="Introduce un número entero ≥ 0 (goles no pueden ser negativos).",
        showInputMessage=True,
        promptTitle="Goles",
        prompt="Introduce el número de goles (entero ≥ 0).",
    )
    dv.sqref = sqref
    return dv

def input_cell(ws, row, col, size=12) -> object:
    """Write and style a user-input goal cell."""
    cell = ws.cell(row=row, column=col)
    cell.fill = fill(INPUT_YELLOW)
    cell.font = Font(name="Calibri", size=size, bold=True, color=DARK_GREY)
    cell.alignment = center()
    cell.border = INPUT_BORDER
    ws.add_data_validation(make_dv(cell.coordinate))
    return cell

def pen_cell(ws, row, col) -> object:
    """Write and style an optional penalty-score cell."""
    cell = ws.cell(row=row, column=col)
    cell.fill = fill(PEN_BG)
    cell.font = Font(name="Calibri", size=9, italic=True, color="8EAADB")
    cell.alignment = center()
    cell.border = PEN_BORDER
    ws.add_data_validation(make_dv(cell.coordinate))
    return cell

# ---------------------------------------------------------------------------
# Tournament data
# ---------------------------------------------------------------------------
GROUPS: dict[str, list[str]] = {
    "A": ["Mexico", "South Korea", "South Africa", "Czech Republic"],
    "B": ["Canada", "Switzerland", "Qatar", "Bosnia & Herzegovina"],
    "C": ["Brazil", "Morocco", "Scotland", "Haiti"],
    "D": ["USA", "Australia", "Paraguay", "Turkey"],
    "E": ["Germany", "Ecuador", "Ivory Coast", "Curaçao"],
    "F": ["Netherlands", "Japan", "Tunisia", "Sweden"],
    "G": ["Belgium", "Iran", "Egypt", "New Zealand"],
    "H": ["Spain", "Uruguay", "Cape Verde", "Saudi Arabia"],
    "I": ["France", "Senegal", "Norway", "Iraq"],
    "J": ["Argentina", "Algeria", "Austria", "Jordan"],
    "K": ["Portugal", "Colombia", "Uzbekistan", "DR Congo"],
    "L": ["England", "Croatia", "Panama", "Ghana"],
}

# (matchday, date, time_et, home, away, venue)
GROUP_MATCHES: dict[str, list[tuple]] = {
    "A": [
        (1, "Jun 11", "3 pm ET",  "Mexico",        "South Africa",     "Mexico City"),
        (1, "Jun 11", "10 pm ET", "South Korea",   "Czech Republic",   "Guadalajara"),
        (2, "Jun 18", "12 pm ET", "Czech Republic","South Africa",     "Atlanta"),
        (2, "Jun 18", "9 pm ET",  "Mexico",        "South Korea",      "Guadalajara"),
        (3, "Jun 24", "9 pm ET",  "Czech Republic","Mexico",           "Mexico City"),
        (3, "Jun 24", "9 pm ET",  "South Africa",  "South Korea",      "Monterrey"),
    ],
    "B": [
        (1, "Jun 12", "3 pm ET",  "Canada",              "Bosnia & Herzegovina", "Toronto"),
        (1, "Jun 13", "3 pm ET",  "Qatar",               "Switzerland",          "San Francisco"),
        (2, "Jun 18", "3 pm ET",  "Switzerland",         "Bosnia & Herzegovina", "Los Angeles"),
        (2, "Jun 18", "6 pm ET",  "Canada",              "Qatar",                "Vancouver"),
        (3, "Jun 24", "3 pm ET",  "Canada",              "Switzerland",          "Vancouver"),
        (3, "Jun 24", "3 pm ET",  "Bosnia & Herzegovina","Qatar",                "Seattle"),
    ],
    "C": [
        (1, "Jun 13", "6 pm ET",  "Brazil",   "Morocco",  "New York/NJ"),
        (1, "Jun 13", "9 pm ET",  "Haiti",    "Scotland", "Boston"),
        (2, "Jun 19", "3 pm ET",  "Scotland", "Morocco",  "Boston"),
        (2, "Jun 19", "9 pm ET",  "Brazil",   "Haiti",    "Philadelphia"),
        (3, "Jun 24", "6 pm ET",  "Scotland", "Brazil",   "Miami"),
        (3, "Jun 24", "6 pm ET",  "Morocco",  "Haiti",    "Atlanta"),
    ],
    "D": [
        (1, "Jun 12", "9 pm ET",  "USA",      "Paraguay", "Los Angeles"),
        (1, "Jun 13", "12 am ET", "Australia","Turkey",   "Vancouver"),
        (2, "Jun 19", "3 pm ET",  "USA",      "Australia","Seattle"),
        (2, "Jun 19", "12 am ET", "Turkey",   "Paraguay", "San Francisco"),
        (3, "Jun 25", "10 pm ET", "Turkey",   "USA",      "Los Angeles"),
        (3, "Jun 25", "10 pm ET", "Paraguay", "Australia","San Francisco"),
    ],
    "E": [
        (1, "Jun 14", "1 pm ET",  "Germany",     "Curaçao",     "Houston"),
        (1, "Jun 14", "7 pm ET",  "Ivory Coast", "Ecuador",     "Philadelphia"),
        (2, "Jun 20", "4 pm ET",  "Germany",     "Ivory Coast", "Toronto"),
        (2, "Jun 20", "8 pm ET",  "Ecuador",     "Curaçao",     "Kansas City"),
        (3, "Jun 25", "4 pm ET",  "Ecuador",     "Germany",     "New York/NJ"),
        (3, "Jun 25", "4 pm ET",  "Curaçao",     "Ivory Coast", "Philadelphia"),
    ],
    "F": [
        (1, "Jun 14", "4 pm ET",  "Netherlands","Japan",       "Dallas"),
        (1, "Jun 14", "10 pm ET", "Sweden",     "Tunisia",     "Monterrey"),
        (2, "Jun 20", "1 pm ET",  "Netherlands","Sweden",      "Houston"),
        (2, "Jun 20", "12 am ET", "Tunisia",    "Japan",       "Monterrey"),
        (3, "Jun 25", "7 pm ET",  "Japan",      "Sweden",      "Dallas"),
        (3, "Jun 25", "7 pm ET",  "Tunisia",    "Netherlands", "Kansas City"),
    ],
    "G": [
        (1, "Jun 15", "3 pm ET",  "Belgium",     "Egypt",       "Seattle"),
        (1, "Jun 15", "9 pm ET",  "Iran",        "New Zealand", "Los Angeles"),
        (2, "Jun 21", "3 pm ET",  "Belgium",     "Iran",        "Los Angeles"),
        (2, "Jun 21", "9 pm ET",  "New Zealand", "Egypt",       "Vancouver"),
        (3, "Jun 26", "11 pm ET", "Egypt",       "Iran",        "Seattle"),
        (3, "Jun 26", "11 pm ET", "New Zealand", "Belgium",     "Vancouver"),
    ],
    "H": [
        (1, "Jun 15", "12 pm ET", "Spain",        "Cape Verde",   "Atlanta"),
        (1, "Jun 15", "6 pm ET",  "Saudi Arabia", "Uruguay",      "Miami"),
        (2, "Jun 21", "12 pm ET", "Spain",        "Saudi Arabia", "Atlanta"),
        (2, "Jun 21", "6 pm ET",  "Uruguay",      "Cape Verde",   "Miami"),
        (3, "Jun 26", "8 pm ET",  "Cape Verde",   "Saudi Arabia", "Houston"),
        (3, "Jun 26", "8 pm ET",  "Uruguay",      "Spain",        "Guadalajara"),
    ],
    "I": [
        (1, "Jun 16", "3 pm ET",  "France",  "Senegal", "New York/NJ"),
        (1, "Jun 16", "6 pm ET",  "Iraq",    "Norway",  "Boston"),
        (2, "Jun 22", "5 pm ET",  "France",  "Iraq",    "Philadelphia"),
        (2, "Jun 22", "8 pm ET",  "Norway",  "Senegal", "New York/NJ"),
        (3, "Jun 26", "3 pm ET",  "Norway",  "France",  "Boston"),
        (3, "Jun 26", "3 pm ET",  "Senegal", "Iraq",    "Toronto"),
    ],
    "J": [
        (1, "Jun 16", "9 pm ET",  "Argentina","Algeria",  "Kansas City"),
        (1, "Jun 16", "12 am ET", "Austria",  "Jordan",   "San Francisco"),
        (2, "Jun 22", "1 pm ET",  "Argentina","Austria",  "Dallas"),
        (2, "Jun 22", "11 pm ET", "Jordan",   "Algeria",  "San Francisco"),
        (3, "Jun 27", "10 pm ET", "Algeria",  "Austria",  "Kansas City"),
        (3, "Jun 27", "10 pm ET", "Jordan",   "Argentina","Dallas"),
    ],
    "K": [
        (1, "Jun 17", "1 pm ET",   "Portugal",  "DR Congo",  "Houston"),
        (1, "Jun 17", "10 pm ET",  "Uzbekistan","Colombia",  "Mexico City"),
        (2, "Jun 23", "1 pm ET",   "Portugal",  "Uzbekistan","Houston"),
        (2, "Jun 23", "10 pm ET",  "DR Congo",  "Colombia",  "Guadalajara"),
        (3, "Jun 27", "7:30 pm ET","Colombia",  "Portugal",  "Miami"),
        (3, "Jun 27", "7:30 pm ET","DR Congo",  "Uzbekistan","Atlanta"),
    ],
    "L": [
        (1, "Jun 17", "4 pm ET",  "England", "Croatia", "Dallas"),
        (1, "Jun 17", "7 pm ET",  "Ghana",   "Panama",  "Toronto"),
        (2, "Jun 23", "4 pm ET",  "England", "Ghana",   "Boston"),
        (2, "Jun 23", "7 pm ET",  "Panama",  "Croatia", "Toronto"),
        (3, "Jun 27", "5 pm ET",  "Panama",  "England", "New York/NJ"),
        (3, "Jun 27", "5 pm ET",  "Croatia", "Ghana",   "Philadelphia"),
    ],
}

REFERENCES = [
    ("FIFA Official",
     "https://www.fifa.com/en/tournaments/mens/worldcup/canadamexicousa2026/articles/match-schedule-fixtures-results-teams-stadiums",
     "Apr 1, 2026",
     "Official match schedule and fixtures"),
    ("FOX Sports",
     "https://www.foxsports.com/stories/soccer/2026-world-cup-schedule-all-games-dates-matchups-how-watch",
     "Apr 1, 2026",
     "Full broadcast schedule with dates, times and venues"),
    ("CNN Español",
     "https://cnnespanol.cnn.com/2026/04/01/deportes/grupos-zonas-copa-mundial-2026-orix",
     "Apr 1, 2026",
     "Complete groups and match calendar (post-playoff)"),
    ("Marca",
     "https://www.marca.com/futbol/mundial/2026/04/01/estan-12-grupos-completos-mundial-falta-irak-bolivia.html",
     "Apr 1, 2026",
     "12 complete groups confirmed; Iraq & DR Congo qualify via inter-confederation playoff"),
    ("Wikipedia – 2026 FIFA World Cup",
     "https://en.wikipedia.org/wiki/2026_FIFA_World_Cup",
     "Apr 1, 2026",
     "Tournament overview, format, host cities, qualification"),
]

# ---------------------------------------------------------------------------
# Tab 1 — Fase de Grupos
# ---------------------------------------------------------------------------
def build_group_stage(wb: Workbook) -> None:
    ws = wb.create_sheet("Fase de Grupos")
    ws.sheet_view.showGridLines = False

    col_widths = {
        1: 4,   # MD
        2: 12,  # Date
        3: 22,  # Home
        4: 6,   # GL
        5: 6,   # GR
        6: 22,  # Away
        7: 20,  # Venue
        8: 3,   # spacer
        9: 4,   # pos
        10: 22, # Team
        11: 5,  # GP
        12: 5,  # W
        13: 5,  # D
        14: 5,  # L
        15: 5,  # GF
        16: 5,  # GA
        17: 5,  # GD
        18: 5,  # Pts
    }
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    row = 1
    group_keys = list(GROUPS.keys())

    for grp in group_keys:
        teams   = GROUPS[grp]
        matches = GROUP_MATCHES[grp]

        # ----------------------------------------------------------------
        # Group header
        # ----------------------------------------------------------------
        easter = ""
        if grp == "A":
            easter = "  ·  ¡Salinator por Chequia! 🇨🇿"
        elif grp == "I":
            easter = "  ·  Allez les bleus! 🇫🇷"

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=18)
        hdr = ws.cell(row=row, column=1, value=f"  GROUP {grp}{easter}")
        hdr.font = Font(name="Calibri", size=14, bold=True,
                        color=WHITE, italic=bool(easter))
        hdr.fill = fill(NAVY)
        hdr.alignment = left_align()
        hdr.border = Border(
            left=_side("medium", GOLD), right=_side("medium", GOLD),
            top=_side("medium", GOLD), bottom=_side("medium", GOLD),
        )
        ws.row_dimensions[row].height = 24
        row += 1

        # ----------------------------------------------------------------
        # Standings table header (cols 9-18)
        # ----------------------------------------------------------------
        standings_hdr_row = row
        for col_off, label in enumerate(
            ["#", "Team", "GP", "W", "D", "L", "GF", "GA", "GD", "Pts"], start=0
        ):
            c = ws.cell(row=row, column=9 + col_off, value=label)
            c.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
            c.fill = fill(DARK_GREY)
            c.alignment = center()
            c.border = THIN_BORDER
        ws.row_dimensions[row].height = 16
        row += 1

        # ----------------------------------------------------------------
        # Standings rows — formulas written after match cells are placed
        # ----------------------------------------------------------------
        standings_start = row
        standing_rows: list[dict] = []

        for team in teams:
            sc(ws, row, 9, "", font=Font(name="Calibri", size=9, bold=True),
               fill_=fill(SILVER), border=THIN_BORDER, align=center())
            sc(ws, row, 10, team,
               font=Font(name="Calibri", size=10, bold=True),
               fill_=fill(WHITE), border=THIN_BORDER, align=left_align())
            for col in range(11, 19):
                sc(ws, row, col, 0,
                   font=Font(name="Calibri", size=10),
                   fill_=fill(WHITE), border=THIN_BORDER, align=center())
            ws.row_dimensions[row].height = 17
            standing_rows.append({"team": team, "row": row})
            row += 1

        row += 1  # spacer between standings and match list

        # ----------------------------------------------------------------
        # Match schedule header (cols 1-7)
        # ----------------------------------------------------------------
        for col_off, label in enumerate(
            ["MD", "Date / Time (ET)", "Home Team", "G", "G", "Away Team", "Venue"],
            start=0,
        ):
            c = ws.cell(row=row, column=1 + col_off, value=label)
            c.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
            c.fill = fill("3D5A80")
            c.alignment = center()
            c.border = THIN_BORDER
        ws.row_dimensions[row].height = 16
        row += 1

        # ----------------------------------------------------------------
        # Match rows
        # ----------------------------------------------------------------
        match_refs: list[dict] = []
        for md, date, time_et, home, away, venue in matches:
            sc(ws, row, 1, md,
               font=Font(name="Calibri", size=9),
               fill_=fill(LIGHT_BLUE), border=THIN_BORDER, align=center())
            sc(ws, row, 2, f"{date} · {time_et}",
               font=Font(name="Calibri", size=8),
               fill_=fill(LIGHT_BLUE), border=THIN_BORDER, align=center())
            sc(ws, row, 3, home,
               font=Font(name="Calibri", size=10, bold=True),
               fill_=fill(LIGHT_BLUE), border=THIN_BORDER, align=left_align())
            hg = input_cell(ws, row, 4)
            ag = input_cell(ws, row, 5)
            sc(ws, row, 6, away,
               font=Font(name="Calibri", size=10, bold=True),
               fill_=fill(LIGHT_BLUE), border=THIN_BORDER, align=left_align())
            sc(ws, row, 7, venue,
               font=Font(name="Calibri", size=8),
               fill_=fill(LIGHT_BLUE), border=THIN_BORDER, align=left_align())
            ws.row_dimensions[row].height = 20
            match_refs.append({
                "home": home, "away": away,
                "hg": hg.coordinate, "ag": ag.coordinate,
            })
            row += 1

        # ----------------------------------------------------------------
        # Build standings formulas
        # ----------------------------------------------------------------
        for sr in standing_rows:
            team = sr["team"]
            srow = sr["row"]
            home_refs = [(m["hg"], m["ag"]) for m in match_refs if m["home"] == team]
            away_refs = [(m["hg"], m["ag"]) for m in match_refs if m["away"] == team]
            all_refs  = home_refs + away_refs

            played  = "+".join(f"IF(AND({h}<>\"\",{a}<>\"\"),1,0)" for h, a in all_refs)
            gf_home = "+".join(f"IF(AND({h}<>\"\",{a}<>\"\"),{h},0)" for h, a in home_refs)
            gf_away = "+".join(f"IF(AND({h}<>\"\",{a}<>\"\"),{a},0)" for h, a in away_refs)
            ga_home = "+".join(f"IF(AND({h}<>\"\",{a}<>\"\"),{a},0)" for h, a in home_refs)
            ga_away = "+".join(f"IF(AND({h}<>\"\",{a}<>\"\"),{h},0)" for h, a in away_refs)
            wins    = (
                "+".join(f"IF(AND({h}<>\"\",{a}<>\"\",{h}>{a}),1,0)" for h, a in home_refs) +
                ("+" if home_refs and away_refs else "") +
                "+".join(f"IF(AND({h}<>\"\",{a}<>\"\",{a}>{h}),1,0)" for h, a in away_refs)
            )
            draws   = "+".join(
                f"IF(AND({h}<>\"\",{a}<>\"\",{h}={a}),1,0)" for h, a in all_refs
            )
            losses  = (
                "+".join(f"IF(AND({h}<>\"\",{a}<>\"\",{h}<{a}),1,0)" for h, a in home_refs) +
                ("+".join(f"IF(AND({h}<>\"\",{a}<>\"\"),0,0)" for h, a in []) or "") +
                ("+" if home_refs and away_refs else "") +
                "+".join(f"IF(AND({h}<>\"\",{a}<>\"\",{a}<{h}),1,0)" for h, a in away_refs)
            )

            gf_formula = "=" + "+".join(filter(None, [gf_home, gf_away])) or "=0"
            ga_formula = "=" + "+".join(filter(None, [ga_home, ga_away])) or "=0"

            ws.cell(row=srow, column=11).value = f"={played}"               # GP
            ws.cell(row=srow, column=12).value = f"={wins}"                 # W
            ws.cell(row=srow, column=13).value = f"={draws}"                # D
            ws.cell(row=srow, column=14).value = f"={losses}"               # L
            ws.cell(row=srow, column=15).value = gf_formula                 # GF
            ws.cell(row=srow, column=16).value = ga_formula                 # GA

            gf_ref = ws.cell(row=srow, column=15).coordinate
            ga_ref = ws.cell(row=srow, column=16).coordinate
            w_ref  = ws.cell(row=srow, column=12).coordinate
            d_ref  = ws.cell(row=srow, column=13).coordinate
            ws.cell(row=srow, column=17).value = f"={gf_ref}-{ga_ref}"     # GD
            ws.cell(row=srow, column=18).value = f"=3*{w_ref}+{d_ref}"     # Pts

        # ----------------------------------------------------------------
        # Conditional formatting on standings (cols 9-18)
        # ----------------------------------------------------------------
        q_range = f"I{standings_start}:R{standings_start + 3}"
        pts_col = "R"

        def _cf_formula(rank: int) -> list[str]:
            base = f"RANK(${pts_col}{standings_start},${pts_col}${standings_start}:${pts_col}${standings_start + 3},0)"
            return [f"{base}={rank}"]

        ws.conditional_formatting.add(q_range, FormulaRule(
            formula=_cf_formula(1),
            fill=PatternFill("solid", fgColor=GREEN_QUAL),
            font=Font(color=GREEN_QUAL_FONT, bold=True),
        ))
        ws.conditional_formatting.add(q_range, FormulaRule(
            formula=_cf_formula(2),
            fill=PatternFill("solid", fgColor="E2EFDA"),
            font=Font(color=GREEN_QUAL_FONT),
        ))
        ws.conditional_formatting.add(q_range, FormulaRule(
            formula=_cf_formula(3),
            fill=PatternFill("solid", fgColor=AMBER_QUAL),
            font=Font(color=AMBER_QUAL_FONT),
        ))
        ws.conditional_formatting.add(q_range, FormulaRule(
            formula=_cf_formula(4),
            fill=PatternFill("solid", fgColor=RED_ELIM),
            font=Font(color=RED_ELIM_FONT),
        ))

        row += 2  # gap between groups

    # ----------------------------------------------------------------
    # Legend
    # ----------------------------------------------------------------
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    sc(ws, row, 1, "  LEGEND",
       font=Font(name="Calibri", size=10, bold=True, color=WHITE),
       fill_=fill(NAVY), border=THIN_BORDER, align=left_align())
    ws.row_dimensions[row].height = 18
    row += 1
    for bg, fc, label in [
        (GREEN_QUAL,  GREEN_QUAL_FONT,  "✅  1st & 2nd place — qualify automatically"),
        ("E2EFDA",    GREEN_QUAL_FONT,  "🟢  2nd place — qualify automatically"),
        (AMBER_QUAL,  AMBER_QUAL_FONT,  "⚠️   3rd place — may qualify as best 3rd"),
        (RED_ELIM,    RED_ELIM_FONT,    "❌  4th place — eliminated"),
        (INPUT_YELLOW, DARK_GREY,       "✏️   Yellow cell — enter goals here (integer ≥ 0)"),
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value=f"  {label}")
        c.font = Font(name="Calibri", size=9, bold=True, color=fc)
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = THIN_BORDER
        c.alignment = left_align()
        ws.row_dimensions[row].height = 16
        row += 1


# ---------------------------------------------------------------------------
# Tab 2 — Bracket
# ---------------------------------------------------------------------------
def build_bracket(wb: Workbook) -> None:
    ws = wb.create_sheet("Bracket")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 70

    # ----------------------------------------------------------------
    # Clean column layout — NO overlapping ranges
    # Each match occupies exactly 4 columns: team_l | gl | gr | team_r
    # Penalty scores go on the row immediately below in the same 4 cols.
    #
    # Left  → center:  R32-L(1-4) R16-L(5-8) QF-L(9-12) SF-L(13-16)
    # Center:           FINAL(17-20)
    # Center → right:  SF-R(21-24) QF-R(25-28) R16-R(29-32) R32-R(33-36)
    # ----------------------------------------------------------------
    TOTAL_COLS = 36

    R32L = (1,  2,  3,  4)
    R16L = (5,  6,  7,  8)
    QFL  = (9,  10, 11, 12)
    SFL  = (13, 14, 15, 16)
    FIN  = (17, 18, 19, 20)
    SFR  = (21, 22, 23, 24)
    QFR  = (25, 26, 27, 28)
    R16R = (29, 30, 31, 32)
    R32R = (33, 34, 35, 36)

    # Row spacing: 4 rows per match slot (label + match + pen + gap)
    SLOT = 4
    BASE = 8  # first R32 match slot

    r32_rows = [BASE + i * SLOT for i in range(8)]
    r16_rows = [(r32_rows[i*2] + r32_rows[i*2+1]) // 2 + 1 for i in range(4)]
    qf_rows  = [(r16_rows[i*2] + r16_rows[i*2+1]) // 2 for i in range(2)]
    sf_row   = (qf_rows[0] + qf_rows[1]) // 2
    final_row = sf_row
    third_row = r32_rows[-1] + SLOT + 3
    max_row   = third_row + SLOT + 2

    # ---- Column widths ------------------------------------------------
    for c in range(1, TOTAL_COLS + 1):
        ltr = get_column_letter(c)
        # team name cols (1st and 4th of each 4-col group)
        if (c - 1) % 4 in (0, 3):
            ws.column_dimensions[ltr].width = 13
        else:
            ws.column_dimensions[ltr].width = 5  # goal cols

    # ---- Paint background --------------------------------------------
    bg = fill(BRACKET_BG)
    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = 18
        for c in range(1, TOTAL_COLS + 1):
            ws.cell(row=r, column=c).fill = bg

    # ---- Title -------------------------------------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=TOTAL_COLS)
    t = ws.cell(row=1, column=1,
                value="🏆  FIFA WORLD CUP 2026  ·  KNOCKOUT BRACKET  🏆")
    t.font = Font(name="Calibri", size=20, bold=True, color=GOLD)
    t.fill = fill(BRACKET_BG)
    t.alignment = center()
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 8

    # ---- Input legend ------------------------------------------------
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=TOTAL_COLS)
    leg = ws.cell(
        row=3, column=1,
        value="  ✏️ Yellow cells = goals (90 min + ET combined)  ·  "
              "Blue-grey rows = Penalty shootout score "
              "(optional — only fill if match is still tied after extra time)",
    )
    leg.font = Font(name="Calibri", size=8, italic=True, color="8EAADB")
    leg.fill = fill(BRACKET_BG)
    leg.alignment = left_align()
    ws.row_dimensions[3].height = 14

    # ---- Round headers -----------------------------------------------
    hdr_row = 6
    ws.row_dimensions[hdr_row].height = 26
    round_specs = [
        (R32L, "ROUND OF 32", "Jun 28 – Jul 3"),
        (R16L, "ROUND OF 16", "Jul 4 – 7"),
        (QFL,  "QUARTER-FINALS", "Jul 9 – 11"),
        (SFL,  "SEMI-FINALS", "Jul 14 – 15"),
        (FIN,  "🏆  FINAL  🏆", "Jul 19 · New York / NJ"),
        (SFR,  "SEMI-FINALS", "Jul 14 – 15"),
        (QFR,  "QUARTER-FINALS", "Jul 9 – 11"),
        (R16R, "ROUND OF 16", "Jul 4 – 7"),
        (R32R, "ROUND OF 32", "Jun 28 – Jul 3"),
    ]
    for cols, round_name, dates in round_specs:
        sc_col, _, _, ec_col = cols
        is_final = "FINAL" in round_name
        ws.merge_cells(start_row=hdr_row, start_column=sc_col,
                       end_row=hdr_row, end_column=ec_col)
        c = ws.cell(row=hdr_row, column=sc_col,
                    value=f"{round_name}\n{dates}")
        c.font = Font(name="Calibri", size=9, bold=True,
                      color=GOLD if is_final else WHITE)
        c.fill = fill(NAVY if is_final else BRACKET_BG)
        c.alignment = center(wrap=True)
        c.border = Border(
            bottom=_side("medium", GOLD if is_final else "3D5A80")
        )

    # ---- Helper: draw one knockout match (match row + optional pen row)
    def draw_match(row, cols, team_l, team_r, date, venue, is_final=False):
        tl_col, gl_col, gr_col, tr_col = cols
        gold_c = GOLD if is_final else "3D5A80"
        tbg    = fill(FINAL_TEAM_BG if is_final else MATCH_BG)

        # Date label row
        lbl_row = row
        ws.merge_cells(start_row=lbl_row, start_column=tl_col,
                       end_row=lbl_row, end_column=tr_col)
        dl = ws.cell(row=lbl_row, column=tl_col,
                     value=f"{date}  ·  {venue}")
        dl.font = Font(name="Calibri", size=7, italic=True,
                       color=GOLD if is_final else "607D8B")
        dl.fill = fill(BRACKET_BG)
        dl.alignment = center()
        ws.row_dimensions[lbl_row].height = 13

        # Match row
        mrow = row + 1
        ws.row_dimensions[mrow].height = 22 if is_final else 20

        tl = ws.cell(row=mrow, column=tl_col, value=team_l or "TBD")
        tl.fill = tbg
        tl.font = Font(name="Calibri", size=12 if is_final else 9,
                       bold=True, color=GOLD if is_final else WHITE)
        tl.alignment = right_align()
        tl.border = Border(
            top=_side("medium" if is_final else "thin", gold_c),
            bottom=_side("medium" if is_final else "thin", gold_c),
            left=_side("medium", gold_c),
        )

        gl = input_cell(ws, mrow, gl_col, size=14 if is_final else 12)
        gr = input_cell(ws, mrow, gr_col, size=14 if is_final else 12)

        tr = ws.cell(row=mrow, column=tr_col, value=team_r or "TBD")
        tr.fill = tbg
        tr.font = Font(name="Calibri", size=12 if is_final else 9,
                       bold=True, color=GOLD if is_final else WHITE)
        tr.alignment = left_align()
        tr.border = Border(
            top=_side("medium" if is_final else "thin", gold_c),
            bottom=_side("medium" if is_final else "thin", gold_c),
            right=_side("medium", gold_c),
        )

        # Penalty row
        prow = row + 2
        ws.row_dimensions[prow].height = 14
        pl_lbl = ws.cell(row=prow, column=tl_col, value="Pen →")
        pl_lbl.font = Font(name="Calibri", size=7, italic=True, color="607D8B")
        pl_lbl.fill = fill(PEN_BG)
        pl_lbl.alignment = right_align()
        pl_lbl.border = PEN_BORDER

        pen_cell(ws, prow, gl_col)
        pen_cell(ws, prow, gr_col)

        pr_lbl = ws.cell(row=prow, column=tr_col, value="← Pen")
        pr_lbl.font = Font(name="Calibri", size=7, italic=True, color="607D8B")
        pr_lbl.fill = fill(PEN_BG)
        pr_lbl.alignment = left_align()
        pr_lbl.border = PEN_BORDER

    # ----------------------------------------------------------------
    # Round of 32 — left side (8 matches)
    # ----------------------------------------------------------------
    r32_left = [
        ("A2",  "B2",    "Jun 28", "Los Angeles"),
        ("C1",  "F2",    "Jun 29", "Houston"),
        ("E1",  "3rd*",  "Jun 29", "Boston"),
        ("F1",  "C2",    "Jun 29", "Monterrey"),
        ("E2",  "I2",    "Jun 30", "Dallas"),
        ("I1",  "3rd*",  "Jun 30", "New York/NJ"),
        ("A1",  "3rd*",  "Jun 30", "Mexico City"),
        ("L1",  "3rd*",  "Jul 1",  "Atlanta"),
    ]
    for i, (tl, tr, date, venue) in enumerate(r32_left):
        draw_match(r32_rows[i], R32L, tl, tr, date, venue)

    # ----------------------------------------------------------------
    # Round of 32 — right side (8 matches)
    # ----------------------------------------------------------------
    r32_right = [
        ("G1",  "3rd*",  "Jul 1",  "Seattle"),
        ("D1",  "3rd*",  "Jul 1",  "San Francisco"),
        ("H1",  "J2",    "Jul 2",  "Los Angeles"),
        ("K2",  "L2",    "Jul 2",  "Toronto"),
        ("B1",  "3rd*",  "Jul 2",  "Vancouver"),
        ("D2",  "G2",    "Jul 3",  "Dallas"),
        ("J1",  "H2",    "Jul 3",  "Miami"),
        ("K1",  "3rd*",  "Jul 3",  "Kansas City"),
    ]
    for i, (tl, tr, date, venue) in enumerate(r32_right):
        draw_match(r32_rows[i], R32R, tl, tr, date, venue)

    # ----------------------------------------------------------------
    # Round of 16
    # ----------------------------------------------------------------
    r16_left = [
        ("W R32 #1", "W R32 #2", "Jul 4", "Houston"),
        ("W R32 #3", "W R32 #4", "Jul 4", "Philadelphia"),
        ("W R32 #5", "W R32 #6", "Jul 5", "New York/NJ"),
        ("W R32 #7", "W R32 #8", "Jul 5", "Mexico City"),
    ]
    r16_right = [
        ("W R32 #9",  "W R32 #10", "Jul 6", "Dallas"),
        ("W R32 #11", "W R32 #12", "Jul 6", "Seattle"),
        ("W R32 #13", "W R32 #14", "Jul 7", "Atlanta"),
        ("W R32 #15", "W R32 #16", "Jul 7", "Vancouver"),
    ]
    for i, (tl, tr, date, venue) in enumerate(r16_left):
        draw_match(r16_rows[i], R16L, tl, tr, date, venue)
    for i, (tl, tr, date, venue) in enumerate(r16_right):
        draw_match(r16_rows[i], R16R, tl, tr, date, venue)

    # ----------------------------------------------------------------
    # Quarterfinals
    # ----------------------------------------------------------------
    qf_left = [
        ("W R16-L1", "W R16-L2", "Jul 9",  "Boston"),
        ("W R16-L3", "W R16-L4", "Jul 10", "Los Angeles"),
    ]
    qf_right = [
        ("W R16-R1", "W R16-R2", "Jul 11", "Miami"),
        ("W R16-R3", "W R16-R4", "Jul 11", "Kansas City"),
    ]
    for i, (tl, tr, date, venue) in enumerate(qf_left):
        draw_match(qf_rows[i], QFL, tl, tr, date, venue)
    for i, (tl, tr, date, venue) in enumerate(qf_right):
        draw_match(qf_rows[i], QFR, tl, tr, date, venue)

    # ----------------------------------------------------------------
    # Semifinals
    # ----------------------------------------------------------------
    draw_match(sf_row, SFL, "W QF-L1", "W QF-L2", "Jul 14", "Dallas")
    draw_match(sf_row, SFR, "W QF-R1", "W QF-R2", "Jul 15", "Atlanta")

    # ----------------------------------------------------------------
    # Final (special styling)
    # ----------------------------------------------------------------
    draw_match(final_row, FIN, "W SF-L", "W SF-R",
               "Jul 19", "New York / NJ · MetLife Stadium", is_final=True)

    # Champion banner below the final
    champ_row = final_row + 3
    ws.merge_cells(start_row=champ_row, start_column=FIN[0],
                   end_row=champ_row, end_column=FIN[3])
    cc = ws.cell(row=champ_row, column=FIN[0], value="🏆  WORLD CHAMPION  🏆")
    cc.font = Font(name="Calibri", size=12, bold=True, color=GOLD)
    cc.fill = fill(NAVY)
    cc.alignment = center()
    cc.border = Border(bottom=_side("medium", GOLD))
    ws.row_dimensions[champ_row].height = 20

    # ----------------------------------------------------------------
    # Third-place match
    # ----------------------------------------------------------------
    ws.merge_cells(start_row=third_row - 1, start_column=FIN[0],
                   end_row=third_row - 1, end_column=FIN[3])
    t3h = ws.cell(row=third_row - 1, column=FIN[0],
                  value="THIRD PLACE MATCH")
    t3h.font = Font(name="Calibri", size=9, bold=True, color=SILVER)
    t3h.fill = fill(BRACKET_BG)
    t3h.alignment = center()
    t3h.border = Border(bottom=_side("thin", SILVER))
    ws.row_dimensions[third_row - 1].height = 14

    draw_match(third_row, FIN, "L SF-L", "L SF-R", "Jul 18", "Miami")


# ---------------------------------------------------------------------------
# Tab 3 — Referencias
# ---------------------------------------------------------------------------
def build_references(wb: Workbook) -> None:
    ws = wb.create_sheet("Referencias")
    ws.sheet_view.showGridLines = False

    for c, w in {1: 24, 2: 72, 3: 14, 4: 50}.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    t = ws.cell(row=1, column=1, value="📚  References / Fuentes")
    t.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = left_align()
    t.border = Border(
        left=_side("medium", GOLD), right=_side("medium", GOLD),
        top=_side("medium", GOLD), bottom=_side("medium", GOLD),
    )
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    sub = ws.cell(row=2, column=1,
                  value="All data sourced from the following public references as of April 1, 2026.")
    sub.font = Font(name="Calibri", size=9, italic=True, color=DARK_GREY)
    sub.fill = fill(SILVER)
    sub.alignment = left_align()
    ws.row_dimensions[2].height = 16

    row = 3
    for col_off, label in enumerate(["Source", "URL", "Accessed", "Description"]):
        c = ws.cell(row=row, column=1 + col_off, value=label)
        c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.fill = fill(DARK_GREY)
        c.alignment = center()
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 18
    row += 1

    for i, (source, url, accessed, desc) in enumerate(REFERENCES):
        row_fill = fill("F2F2F2" if i % 2 == 0 else WHITE)
        for col_off, val in enumerate([source, url, accessed, desc]):
            col = 1 + col_off
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(
                name="Calibri", size=9,
                color="0563C1" if col == 2 else DARK_GREY,
                underline="single" if col == 2 else None,
            )
            c.fill = row_fill
            c.border = THIN_BORDER
            c.alignment = left_align(wrap=(col in (2, 4)))
        ws.row_dimensions[row].height = 32
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    note = ws.cell(
        row=row, column=1,
        value=(
            "Tournament format: 48 teams · 12 groups of 4 · "
            "Top 2 per group + 8 best 3rd-place finishers → Round of 32 · "
            "104 total matches · June 11 – July 19, 2026 · "
            "Host cities: USA, Canada, Mexico"
        ),
    )
    note.font = Font(name="Calibri", size=9, italic=True, color=DARK_GREY)
    note.fill = fill("FFF9C4")
    note.alignment = left_align(wrap=True)
    note.border = THIN_BORDER
    ws.row_dimensions[row].height = 36

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    gen = ws.cell(
        row=row, column=1,
        value="Generated by generate_quiniela.py  ·  MIT License",
    )
    gen.font = Font(name="Calibri", size=8, italic=True, color="888888")
    gen.alignment = left_align()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    wb = Workbook()
    del wb["Sheet"]

    print("Building group stage tab…")
    build_group_stage(wb)

    print("Building bracket tab…")
    build_bracket(wb)

    print("Building references tab…")
    build_references(wb)

    wb.active = wb["Fase de Grupos"]
    wb.save(OUTPUT_FILE)
    print(f"✅  Saved → {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
